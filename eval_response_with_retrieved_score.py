import argparse
import os

import numpy as np
import torch
from numpy import linalg
from tensorboardX import SummaryWriter
from torch.optim.adamw import AdamW
from torch.utils.data import DataLoader, Dataset, RandomSampler
from tqdm import tqdm
from transformers import BertConfig, BertModel, BertTokenizer
from string import ascii_uppercase
from openpyxl import Workbook
from openpyxl.styles import Font
import random
from pprint import pprint
from get_dataset import get_dd_corpus, get_zhao_dataset, get_persona_corpus
from utils import (
    dump_config,
    get_correlation,
    draw_scatter_plot,
    load_model,
    save_model,
    softmax_2d,
    set_random_seed,
    get_usencoder,
    write_summary,
)
import numpy as np

from nlgeval import NLGEval
from sklearn.metrics.pairwise import cosine_similarity
from tqdm import tqdm
from bert_score import score as bertscore
from nltk.translate.bleu_score import sentence_bleu
from matplotlib import pyplot as plt
import pickle
from retrieve_with_bi_encoder import make_annotated_dataset

import argparse

NAMEMAP = {'avg':'Emb. Average','grd':"Emb. Greedy",'ext':"Emb. Extrema",'meteor':"METEOR",'rouge':"ROUGE",'bleu1':'BLEU1','bleu2':'BLEU2','use':'USE'}


def main(args):
    with open(args.similarity_fname, "rb") as f:
        similarity_data = pickle.load(f)

    human_score = similarity_data["human"]
    candidate_size_list = [1,50,100,200,300,400,500]

    correlation_result = {}
    prediction_result = {"human": human_score}
    for metric_name, metric_result in similarity_data.items():
        if metric_name == "human":
            continue
        correlation_result[metric_name] = {}
        correlation_result[metric_name]['max'] = {}
        correlation_result[metric_name]['mean'] = {}
        correlation_result[metric_name]['min'] = {}
        correlation_result[metric_name]['random'] = {}

        similarity_list, reference_score = metric_result["similarity"], metric_result["ref"]
        # Reference Only
        ref_only_correlation = get_correlation(human_score, reference_score)
        correlation_result[metric_name]["refonly"] = ref_only_correlation
        prediction_result[metric_name + "-ref"] = reference_score

        for candidate_size in candidate_size_list:
            # Ours
            max_scores = np.max(np.array([el[:candidate_size] for el in similarity_list]), 1)
            retrieved_max_correlation = get_correlation(human_score, max_scores)
            mean_scores = np.mean(np.array([el[:candidate_size] for el in similarity_list]), 1)
            retrieved_mean_correlation = get_correlation(human_score, mean_scores)
            min_scores = np.min(np.array([el[:candidate_size] for el in similarity_list]), 1)
            retrieved_min_correlation = get_correlation(human_score, min_scores)
            random_scores = np.array([random.sample(el[:candidate_size],1)[0] for el in similarity_list])
            retrieved_random_correlation = get_correlation(human_score, random_scores)

            correlation_result[metric_name]['max'][candidate_size] = retrieved_max_correlation
            correlation_result[metric_name]['mean'][candidate_size] = retrieved_mean_correlation            
            correlation_result[metric_name]['min'][candidate_size] = retrieved_min_correlation
            correlation_result[metric_name]['random'][candidate_size] = retrieved_random_correlation

            prediction_result[metric_name + f"-{candidate_size}-max"] = max_scores
            prediction_result[metric_name + f"-{candidate_size}-mean"] = mean_scores
            prediction_result[metric_name + f"-{candidate_size}-min"] = min_scores
            prediction_result[metric_name + f"-{candidate_size}-random"] = random_scores

    # Excel
    if True:
        wb = Workbook()
        ws = wb.active

        row_count = 2
        ws["A1"] = "Metric"
        ws["B1"] = "# candidate"
        ws["C1"] = "aggregate func."
        ws["D1"] = "pearson"
        ws["E1"] = "spearman"

        for metric_name, aggregation_item in correlation_result.items():
            for aggregation_name, aggregation_result in aggregation_item.items():
                if aggregation_name == 'refonly':
                    assert len(aggregation_result) == 2
                    ws['A'+str(row_count)] = metric_name
                    ws['A'+str(row_count)].font = Font(bold=True)
                    ws['D'+str(row_count)] = aggregation_result[0]
                    ws['D'+str(row_count)].font = Font(bold=True)
                    ws['E'+str(row_count)] = aggregation_result[1]
                    ws['E'+str(row_count)].font = Font(bold=True)
                    row_count += 1
                    continue
                for candidate_size,candidate_result in aggregation_result.items():
                    ws["A" + str(row_count)] = metric_name
                    ws["B" + str(row_count)] = candidate_size
                    ws["C" + str(row_count)] = aggregation_name
                    assert len(candidate_result) == 2
                    ws["D" + str(row_count)] = candidate_result[0]
                    ws["E" + str(row_count)] = candidate_result[1]
                    row_count += 1

        os.makedirs('excel_result',exist_ok=True)
        wb.save("./excel_result/{}-result.xlsx".format(args.eval_set))

    # Plot with different metrics with a varying the number of candidates
    if False:
        score_index = 0 # 0 for pearson, 1 for spearman
        assert len(correlation_result) == 8
        
        for score_index in range(2):
            count = 0
            fig, axs = plt.subplots(2,4,sharex=True,sharey=True)
            
            for metric_name, item in correlation_result.items():
                x = count // 4
                y = count % 4
                ref_only_score = item['refonly'][score_index]
                scores_with_vary = [item['max'][candi_size][score_index] for candi_size in candidate_size_list]
                axs[x, y].axhline(y=ref_only_score,color='r')
                axs[x, y].plot(candidate_size_list, scores_with_vary)
                axs[x, y].set_title(NAMEMAP[metric_name])
                axs[x, y].set_ylim([0.0, 0.8])
                axs[x, y].set_xticks(candidate_size_list)
                axs[x, y].tick_params(labelsize=6)
                axs[x, y].tick_params(axis='x',labelrotation=45)
                
                count += 1
            
            plt.tight_layout()
            os.makedirs('image',exist_ok=True)
            plt.savefig("./image/{}-plot-{}.png".format(args.eval_set,['pearson','spearman'][score_index]), dpi=300)
            plt.close()

    # Plot with different aggregation methods
    if False:
        score_index = 0 # 0 for pearson, 1 for spearman
        assert len(correlation_result) == 8
        
        for score_index in range(2):
            count = 0
            fig, axs = plt.subplots(2,4,sharex=True,sharey=True)
            
            for metric_name, item in correlation_result.items():
                x = count // 4
                y = count % 4
                ref_only_score = item['refonly'][score_index]
                max_scores_with_vary = [item['max'][candi_size][score_index] for candi_size in candidate_size_list]
                mean_scores_with_vary = [item['mean'][candi_size][score_index] for candi_size in candidate_size_list]
                min_scores_with_vary = [item['min'][candi_size][score_index] for candi_size in candidate_size_list]
                random_scores_with_vary = [item['random'][candi_size][score_index] for candi_size in candidate_size_list]
                axs[x, y].axhline(y=ref_only_score,color='r')
                axs[x, y].plot(candidate_size_list, max_scores_with_vary,label='max')
                axs[x, y].plot(candidate_size_list, mean_scores_with_vary,label='mean')
                axs[x, y].plot(candidate_size_list, min_scores_with_vary,label='min')
                axs[x, y].plot(candidate_size_list, random_scores_with_vary,label='random')
                axs[x, y].set_title(NAMEMAP[metric_name])
                axs[x, y].set_ylim([0.0, 0.8])
                axs[x, y].set_xticks(candidate_size_list)
                axs[x, y].tick_params(labelsize=6)
                axs[x, y].tick_params(axis='x',labelrotation=45)
                
                count += 1
            plt.legend()
            plt.tight_layout()
            os.makedirs('image',exist_ok=True)
            plt.savefig("./image/agg_func-{}-plot-{}.png".format(args.eval_set,['pearson','spearman'][score_index]), dpi=300)
            plt.close()
            
    # Scatter Plot
    if False:
        print(prediction_result.keys())
        input()

        for metric_name, metric_prediction in prediction_result.items():
            draw_scatter_plot(
                metric_name,
                prediction_result["human"],
                metric_prediction,
                "./scatter/" + "{}.png".format(metric_name),
            )




if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process some integers.")
    parser.add_argument("--ir_method", type=str, default="bi_encoder", choices=["bi_encoder", "bm25"])
    parser.add_argument("--eval_set", type=str, default="dd", choices=["dd", "persona"])
    
    parser.add_argument("--max_candidate", type=int, default=500)
    args = parser.parse_args()

    args.similarity_fname = "./ours_data/{}-{}-{}-similarity.pck".format(
        args.eval_set, args.ir_method, args.max_candidate
    )

    main(args)